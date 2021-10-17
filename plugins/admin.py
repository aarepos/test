from re import search
from pyrogram import Client, filters
from tinydb import TinyDB, Query
from . import config, pay, main_menu
import os
import time
import openpyxl
import xlsxwriter
import datetime
from pyrogram.types import (
    InlineKeyboardMarkup,
    InlineKeyboardButton
)


ch = TinyDB('channels.json')
db = TinyDB('user.json')
admin = Query()
user = Query()


@Client.on_message(filters.command('set') & filters.user(config.admin))
def set_channels(client, message):

    if len(message.command) == 3:

        uid = message.from_user.id
        pid = message.command[1]
        cid = message.command[2]
        mode = 'all'

    elif len(message.command) == 4:
        uid = message.from_user.id
        pid = message.command[1]
        cid = message.command[2]
        mode = message.command[3]

    bid = client.get_me().id

    try:

        result = client.get_chat_member(cid, bid)
        if result.status == 'administrator':

            search_for_channel = ch.search(admin.id == pid)
            if len(search_for_channel) != 0:
                ch.update({'cid': cid, 'mode': mode}, admin.id == pid)
                state = 'UPDATED'
            else:
                ch.insert({'id': pid, 'cid': cid, 'mode': mode})
                state = 'INSERTED'

            # all_users = TinyDB('user.json').all()
            # for person in all_users:

            #     user_type = person['user_type']
                
            #     if user_type in [2, 3]:

            #         try:
            #             invite_link = client.create_chat_invite_link(
            #                 chat_id=cid,
            #                 member_limit=1
            #             )['invite_link']
            #         except Exception as e:
            #             continue


            #         client.send_message(
            #             chat_id = person['id'],
            #             text = f'channel updated! [join now]({invite_link})',
            #             parse_mode = 'markdown'
            #         )

            #     if user_type == 1:

            #         try:
            #             invite_link = client.create_chat_invite_link(
            #                 chat_id=cid,
            #                 member_limit=1
            #             )['invite_link']
            #         except Exception as e:
            #             continue

            #         if mode == 'bronze':
            #             client.send_message(
            #                 chat_id = person['id'],
            #                 text = f'channel updated! [join now]({invite_link})',
            #                 parse_mode = 'markdown'
            #             )
            #         else:

            #             search_for_host = TinyDB('user.json').search(user.host == person['id'])
            #             count = len(search_for_host) / 2

            #             if pid <= count:
            #                 client.send_message(
            #                     chat_id = person['id'],
            #                     text = f'channel updated! [join now]({invite_link})',
            #                     parse_mode = 'markdown'
            #                 )




            #         client.send_message(
            #             chat_id = person['id'],
            #             text = f'channel updated! [join now]({invite_link})'
            #         )

            #     if user_type == 0:

            #         try:
            #             invite_link = client.create_chat_invite_link(
            #                 chat_id=cid,
            #                 member_limit=1
            #             )['invite_link']
            #         except Exception as e:
            #             continue
                    
            #         search_for_host = TinyDB('user.json').search(user.host == person['id'])
            #         count = len(search_for_host) / 2

            #         if pid <= count:
            #             client.send_message(
            #                 chat_id = person['id'],
            #                 text = f'channel updated! [join now]({invite_link})',
            #                 parse_mode = 'markdown'
            #             )

            message.reply_text(f'OK, {state}')
            
        else:
            message.reply_text(f'please admin the robot in channel ({cid})')

    except Exception as e:
        message.reply_text(f'please admin the robot in channel ({cid})')


@Client.on_message(filters.command('ban') & filters.user(config.admin))
def ban(client, message):

    if len(message.command) == 1:
        return

    for person in message.command:

        if person == '/ban' or person == config.admin:
            continue

        try:
            db.update({'status': 'block'}, user.id == int(person))
        except Exception as e:
            continue

    message.reply_text('OK')


@Client.on_message(filters.command('unban') & filters.user(config.admin))
def unban(client, message):

    if len(message.command) == 1:
        return

    for person in message.command:

        if person == '/unban' or person == config.admin:
            continue

        try:
            db.update({'status': 0}, user.id == int(person))
        except Exception as e:
            continue

    message.reply_text('OK')


@Client.on_message(filters.command('chtype') & filters.user(config.admin))
def chtype(client, message):

    if len(message.command) != 3:
        return

    user_id = int(message.command[1])
    user_type_to = message.command[2]

    if user_type_to == 'simple':
        user_type_to = 'None'

    if user_type_to not in config.user_types:
        message.reply_text(f'Error: [{user_type_to}]')
        return

    try:

        client.send_message(
            chat_id=user_id,
            text=f'Upgrade to {user_type_to}'
        )

        if user_type_to == 'None':

            db.update({'expire': None, 'user_type': 0,
                      'search_per_day': 1}, user.id == int(user_id))
            
            message.reply_text('OK')
            main_menu.show_menu(client, user_id)
            return

        now = int(time.time())
        get_period = config.period[user_type_to]
        expire = 'unlimited' if get_period == 0 else now + (get_period * 86400)

        db.update({'expire': expire, 'user_type': int(config.user_types.index(user_type_to)),
                  'search_per_day': config.user_search[user_type_to]}, user.id == int(user_id))

        if user_type_to == 'bronze':
            pay.send_private_channels(client, int(user_id), mode='bronze')

        else:
            pay.send_private_channels(client, int(user_id))
        
        message.reply_text('OK')

    except Exception as e:
        message.reply_text('Error! /chtype USERID USERTYPE')


@Client.on_message(filters.command('chlim') & filters.user(config.admin))
def chlim(client, message):

    if len(message.command) != 3:
        return

    person = int(message.command[1])
    lim = int(message.command[2])

    try:
        user_type = db.search(user.id == person)[0]['user_type']

        if user_type != 0:
            message.reply_text("This user is vip!")
            return

        db.update({'search_per_day': lim}, user.id == person)
        message.reply_text('OK')
    except Exception as e:
        message.reply_text('Error!')


@Client.on_message(filters.command('sendfor') & filters.user(config.admin))
def sendfor(client, message):

    if len(message.command) != 2:
        return

    mid = message.reply_to_message.message_id

    user_type = message.command[1]

    if user_type == 'all':

        for person in db.all():
            if person['id'] == config.admin:
                continue

            client.copy_message(
                chat_id=person['id'],
                from_chat_id=config.admin,
                message_id=mid
            )

        message.reply_text('Sent')
        return

    try:

        if user_type == 'simple':
            user_type = 'None'

        users = TinyDB('user.json').search(user.user_type == int(config.user_types.index(user_type)))
        for person in users:
            if person['id'] == config.admin:
                continue
            client.copy_message(
                chat_id=person['id'],
                from_chat_id=config.admin,
                message_id=mid
            )

        message.reply_text('Sent')

    except Exception as e:
        message.reply_text('Error!')


@Client.on_message(filters.command('sendexel') & filters.user(config.admin))
def send_with_exel(client, message):

    if len(message.command) != 2:
        return

    text = message.command[1]
    path = f'files/{text}.xlsx'

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    rows = row = sheet_obj.max_row

    for i in range(1, rows + 1):

        try:

            user_id = sheet_obj.cell(row=i, column=1).value
            msg = sheet_obj.cell(row=i, column=2).value
            client.send_message(
                chat_id=int(user_id),
                text=str(msg)
            )

        except Exception as e:
            continue


@Client.on_message(filters.command('q') & filters.user(config.admin))
def question(client, message):

    if len(message.command) != 2:
        return

    text = message.command[1]
    path = f'files/{text}.xlsx'

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    rows = row = sheet_obj.max_row

    cuser = 0
    for i in range(1, rows + 1):

        try:

            reply_button = [
                [
                    InlineKeyboardButton(
                        'Yes',
                        # [1] => row, [2] => file
                        callback_data=f'yes1-{i}-{text}'
                    ),
                    InlineKeyboardButton(
                        'No',
                        callback_data=f'no1-{i}-{text}'
                    ),
                ]
            ]

            user_id = sheet_obj.cell(row=i, column=1).value
            if user_id == config.admin:
                continue
            client.send_message(
                chat_id=int(user_id),
                text=config.question_text,
                reply_markup=InlineKeyboardMarkup(reply_button)
            )
            cuser += 1

        except Exception as e:
            continue

    message.reply_text(f'Sent for {cuser} users.')


@Client.on_message(filters.command('a') & filters.user(config.admin))
def answer(client, message):

    if len(message.command) != 2:
        return

    exel_file_name = message.command[1]
    exel_data = TinyDB(f'files/answers/{exel_file_name}.json')

    workbook = xlsxwriter.Workbook(f'files/answers/{exel_file_name}.xlsx')
    worksheet = workbook.add_worksheet()

    data_list = []

    for person in exel_data.all():

        uid = int(person['sender'])
        ans = str(person['answer'])

        user_info = db.search(user.id == uid)[0]
        username = user_info['username']
        fname = user_info['fname']
        host = user_info['host']
        host = 'nobody' if host == None else host
        user_type = user_info['user_type']
        search_per_day = user_info['search_per_day']
        search_count = user_info['search_count']
        expire = user_info['expire']
        expire = expire / 86400 if expire != None else 0

        user_data = [
            uid,
            ans,
            username,
            fname,
            host,
            user_type,
            search_per_day,
            search_count,
            expire
        ]

        data_list.append(user_data)

    data_tuple = tuple(data_list)

    worksheet.write(0, 0, 'user id')
    worksheet.write(0, 1, 'answer')
    worksheet.write(0, 2, 'username')
    worksheet.write(0, 3, 'first name')
    worksheet.write(0, 4, 'host')
    worksheet.write(0, 5, 'user type')
    worksheet.write(0, 6, 'search per day')
    worksheet.write(0, 7, 'search count')
    worksheet.write(0, 8, 'expire')

    row = 1
    col = 0
    for item in (data_tuple):
        worksheet.write(row, col,     item[0])
        worksheet.write(row, col + 1, item[1])
        worksheet.write(row, col + 2, item[2])
        worksheet.write(row, col + 3, item[3])
        worksheet.write(row, col + 4, item[4])
        worksheet.write(row, col + 5, item[5])
        worksheet.write(row, col + 6, item[6])
        worksheet.write(row, col + 7, item[7])
        worksheet.write(row, col + 8, item[8])
        row += 1

    workbook.close()

    client.send_document(
        chat_id=config.admin,
        document=f'files/answers/{exel_file_name}.xlsx'
    )

    os.remove(f'files/answers/{exel_file_name}.xlsx')


@Client.on_message(filters.command('info') & filters.user(config.admin))
def info(client, message):

    if len(message.command) != 2:
        return

    user_id = int(message.command[1])
    user_info = db.search(user.id == user_id)

    if len(user_info) == 0:

        message.reply_text('User not found')
    else:
        inf = user_info[0]
        user_information = 'USER INFORMATION \n\n'
        user_information += 'username: @' + inf['username'] + '\n'
        user_information += 'first name: ' + str(inf['fname']) + '\n'
        user_information += 'host: ' + str(inf['host']) + '\n'
        user_information += 'user type: ' + \
            str(config.user_types[inf['user_type']]) + '\n'
        user_information += 'search count: ' + str(inf['search_count']) + '\n'
        expire = 0 if inf['expire'] == None else float(inf['expire']) / 86400
        user_information += 'expire(day): ' + str(expire) + '\n'
        search_for_subs = db.search(user.host == user_id)
        user_information += 'subcategories: ' + \
            str(len(search_for_subs)) + '\n'

        message.reply_text(user_information)
